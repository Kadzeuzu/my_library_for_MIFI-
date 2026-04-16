import json
import re
from datetime import datetime
import openpyxl

MONTHS = {1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
          7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'}

def format_date(date_str):
    try:
        dt = datetime.strptime(date_str.strip(), '%d.%m.%Y')
        return f"{dt.day} {MONTHS[dt.month]}"
    except:
        return date_str

def get_first_nonempty_line(cell_value):
    if not cell_value:
        return "", -1
    lines = str(cell_value).split('\n')
    for i, line in enumerate(lines):
        if line.strip():
            return line.strip(), i
    return "", -1

def parse_schedule(file_path, target_group="Б24-142", week="верхняя"):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.worksheets[1]  # второй лист

    subj_col, teach_col, room_col = 7, 8, 9
    other_subj_col, other_teach_col, other_room_col = 4, 5, 6

    header_row = None
    for row in range(1, min(30, sheet.max_row)):
        row_vals = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[row]]
        if any("д/н" in v for v in row_vals) and any("время" in v for v in row_vals):
            header_row = row
            break
    if header_row is None:
        raise ValueError("Заголовок не найден")
    print(f"Заголовок найден на строке {header_row}")

    schedule = []
    current_day = None
    current_date = None

    for row in range(header_row + 2, sheet.max_row + 1):
        first_cell = sheet.cell(row=row, column=1).value
        if first_cell and "Внесено" in str(first_cell):
            print(f"Стоп на строке {row} (Внесено)")
            break

        day_cell = sheet.cell(row=row, column=1).value
        if day_cell:
            day_match = re.match(r"(Понедельник|Вторник|Среда|Четверг|Пятница|Суббота)", str(day_cell))
            if day_match:
                current_day = day_match.group(1)
                date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', str(day_cell))
                current_date = format_date(date_match.group(1)) if date_match else ""
                # print(f"День: {current_day}, дата: {current_date}")

        if not current_day:
            continue

        time_cell = sheet.cell(row=row, column=3).value
        if not time_cell:
            continue
        time_str = str(time_cell).strip()
        if not re.match(r'\d{1,2}\.\d{2}-\d{1,2}\.\d{2}', time_str):
            continue

        subject_cell = sheet.cell(row=row, column=subj_col).value
        teacher_cell = sheet.cell(row=row, column=teach_col).value
        room_cell = sheet.cell(row=row, column=room_col).value

        subject, subj_idx = get_first_nonempty_line(subject_cell)
        teacher, _ = get_first_nonempty_line(teacher_cell)
        room, _ = get_first_nonempty_line(room_cell)

        # Если нет предмета, пытаемся взять из другой группы
        if not subject:
            other_subject_cell = sheet.cell(row=row, column=other_subj_col).value
            other_teacher_cell = sheet.cell(row=row, column=other_teach_col).value
            other_room_cell = sheet.cell(row=row, column=other_room_col).value
            other_subject, other_idx = get_first_nonempty_line(other_subject_cell)
            if other_subject:
                subject = other_subject
                subj_idx = other_idx
                if not teacher:
                    teacher, _ = get_first_nonempty_line(other_teacher_cell)
                if not room:
                    room, _ = get_first_nonempty_line(other_room_cell)

        if not subject:
            continue

        # Отладочная печать для проблемных строк
        if current_day == "Четверг" and "14.10-15.30" in time_str:
            print(f"Четверг 4 пара: subject='{subject}', idx={subj_idx}, teacher='{teacher}', room='{room}'")
        if current_day == "Суббота" and "10.30-11.50" in time_str:
            print(f"Суббота 2 пара: subject='{subject}', idx={subj_idx}, teacher='{teacher}', room='{room}'")

        if week == "верхняя" and subj_idx != 0:
            print(f"Пропуск (нижняя неделя): {current_day} {time_str} '{subject}' (idx={subj_idx})")
            continue

        if "внесено" in subject.lower() or "согласовано" in subject.lower():
            continue

        schedule.append({
            "day": current_day,
            "date": current_date,
            "time": time_str,
            "name": subject,
            "teacher": teacher,
            "room": room
        })

    def time_to_minutes(t):
        try:
            start = t.split('-')[0]
            h, m = map(int, start.split('.'))
            return h * 60 + m
        except:
            return 0

    days_dict = {}
    for item in schedule:
        day = item["day"]
        if day not in days_dict:
            days_dict[day] = []
        days_dict[day].append(item)
    for day in days_dict:
        days_dict[day].sort(key=lambda x: time_to_minutes(x["time"]))

    day_order = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
    result_days = []
    for day in day_order:
        if day in days_dict:
            date = days_dict[day][0]["date"]
            result_days.append({
                "day": day,
                "date": date,
                "lessons": [{"time": l["time"], "name": l["name"], "teacher": l["teacher"], "room": l["room"]} for l in days_dict[day]]
            })

    return {
        "week_title": f"{week} неделя",
        "original_link": file_path,
        "days": result_days
    }

if name == "main":
    import sys
    file_path = sys.argv[1] if len(sys.argv) > 1 else "schedule.xlsx"
    data = parse_schedule(file_path, target_group="Б24-142", week="верхняя")
    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)
    print("Готово! data.json обновлён.")
